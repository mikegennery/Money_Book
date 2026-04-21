'''
CAVC Coding Academy Front End
Michael Gennery – 764946

Money Book
Personal Expense Tracker

Version 1.0

Date 11/04/2026

'''


# Import statements
import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox
import pandas as pd
import tkinter.ttk as ttk
import datetime as dt
from openpyxl import load_workbook
import matplotlib.pyplot as plt


# FUNCTIONS


# FUNCTION - Entry Form

def entryForm():

    # FUNCTION - Save Record (sub function of entryForm)

    def saveRecord():
            # Obtain Values from Entry Fields
            description = descriptionEntry.get().strip()
            category = categoryEntry.get().strip()
            # Ensure a valid currency amount is entered
            validAmount = True
            try:
                amount = float(amountEntry.get())
            except ValueError:
                validAmount = False
                amount = 0

            # VALIDATION - Ensure data entered is correct and valid
            if not validAmount or amount < 0.01 or amount > 999.99:
                messagebox.showinfo("Warning!","Enter a valid currency amount between 0.1 and 999.99")
            elif description == "":
                messagebox.showinfo("Warning!","Description cannot be blank")
            elif description == "" or category not in categoriesList:
                messagebox.showinfo("Warning!","Please select a valid category from the drop down list")
            else:
                # Save record to file and display records table
                # https://www.datacamp.com/tutorial/python-excel-tutorial
                pet.append([today, description, category, amount])
                petWorkBook.save("Money_Book.xlsx")
                todayText = str(today)[8:10] + '/' + str(today)[5:7] + '/' + str(today)[:4] # Format the date DD/MM/YYYY
                amountText = str(f"£{amount:.2f}") # Format amount as currency
                recordsTable.insert("", "end", values=(todayText, description, category, amountText))
                messagebox.showinfo("Money Book",description + ' ( ' + category + ' ): £' + str(f"{amount:.2f}") + '\n' + 'Saved')

    entryFormWindow = tk.Tk()
    entryFormWindow.title("Money Book")
    entryFormWindow.geometry("600x400")

    entryFormFrame = tk.Frame(entryFormWindow)
    entryFormFrame.pack(pady=10)

    tk.Label(entryFormFrame, text="Entry Form", font=(font, heading1)).grid(row=0, column=0, padx=5)

    # Description Entry Field
    tk.Label(entryFormFrame, text="Description", width=20, font=(font, textSize)).grid(row=1, column=0, padx=5)
    descriptionEntry = tk.Entry(entryFormFrame)
    descriptionEntry.grid(row=1, column=1, padx=5)
    
    # Category Entry Field
    categoryValue = tk.StringVar()
    tk.Label(entryFormFrame, text="Category", font=(font, textSize)).grid(row=2, column=0, padx=5)
    categoryEntry = ttk.Combobox(entryFormFrame, width=17, textvariable=categoryValue)
    categoryEntry.grid(row=2, column=1, padx=5)

    # List of Categories
    categoryEntry['values'] = categoriesList

    # Amount Description
    tk.Label(entryFormFrame, text="Amount", font=(font, textSize)).grid(row=3, column=0, padx=5)
    amountEntry = tk.Entry(entryFormFrame, width=20)
    amountEntry.grid(row=3, column=1, padx=5)

    # Save Button
    saveButton = tk.Button(entryFormFrame, text="SAVE", font=(font, textSize), command=saveRecord)
    saveButton.grid(row=4, column=0, padx=5)

    # Display list of expense transactions

    # Obtain List of headings
    columns = []
    for column in pet[1]:
        columns.append(column.value)
    
    # Create table to display transactions
    recordsTable = ttk.Treeview(entryFormWindow, columns=columns, show='headings')
    # Set Headings and columns
    for value in columns:
        recordsTable.heading(value, text=value)
        recordsTable.column(value, width=150, anchor=tk.CENTER)
    
    # Insert records from database into the table on the screen
    for record in pet.iter_rows(min_row=2, values_only=TRUE):
        # Create a tuple with fields properly formatted for date (DD/MM/YYYY) and amount e.g. £9.99
        formattedRecord = (str(record[0])[8:10]) + '/' + str(record[0])[5:7] + '/' + str(record[0])[:4],record[1],record[2],str(f"£{record[3]:.2f}")
        recordsTable.insert("", "end", values=formattedRecord)

    recordsTable.bind("<<TreeViewSelect>>")
    recordsTable.pack(fill=BOTH, expand=TRUE, padx=10, pady=10)

    entryFormWindow.mainloop()


# FUNCTION - Return 2D array with transaction records

def obtainRecords():
    recordsTable = []
    # Go through each record in the database
    for record in pet:
        recordArray = []
        # Go through each field in the record and add to array
        for field in record:
            recordArray.append(field.value)
        # Add record array to table array
        recordsTable.append(recordArray)
    # Return completed array to requesting function
    return recordsTable


# GRAPHS

# record[0] - Date
# record[1] - Description
# record[2] - Category
# record[3] - Amount


# FUNCTION - Categories
# Create and display a pie chart for each category

def categories():    
    tableA = createTableA() # Amounts for each category

    # Create a pie chart with tableA
    plt.figure(figsize=(5, 5))
    plt.pie(list(tableA.values()), labels=list(tableA.keys()), autopct="%1.1f%%")
    plt.title('Categories')
    plt.show()


# FUNCTION - Monthly Balances - Create and display a bar chart for monthly amounts
# https://www.datacamp.com/tutorial/python-bar-plot

def monthlyBalances():
    tableB = createTableB() # Amounts for each month

    # Create a bar chart with tableA
    plt.figure(figsize=(5, 5))
    plt.bar(tableB.keys(), tableB.values())
    plt.xlabel('Month')
    plt.ylabel('Amount £')
    plt.title('Monthly Balances')
    for i, value in enumerate(tableB.values()):
        plt.text(i, value, str(f"£{value:.2f}"), ha='center')
    plt.show()


# FUNCTION - Statistics

def statistics():
    statisticsWindow = tk.Tk()
    statisticsWindow.title("Money Book")
    statisticsWindow.geometry("600x400")

    tk.Label(statisticsWindow, text="Statistics", font=(font, heading3)).pack()

    tableB = createTableB() # Amounts for each month
    tableC = createTableC() # Amounts for each category for each month

    # Extract the last three months
    tableCList = list(tableC.keys())

    for month in tableCList[-2:]:
        # Display totals for each month from tableB
        tk.Label(statisticsWindow, text=calendar[month[5:7]] + ' ' + month[0:4] + ': £' + str(f"{tableB[month]:.2f}"), font=(font, heading3)).pack()
        
        for category in tableC[month]:
            # Display totals for each category in each month
            tk.Label(statisticsWindow, text=category + ': £' + str(f"{tableC[month][category]:.2f}"), font=(font, textSize)).pack()
            
    statisticsWindow.mainloop()


# FUNCTION - Create Table A
# Amounts for each category

def createTableA():
    # Obtain 2D array with transaction records
    recordsTable = obtainRecords()
    tableA = {}

    # If category is in table A, increment by 1
    # otherwise create a new record
    for record in recordsTable:
        if record[2] != 'CATEGORY': # Don't count the heading
            # Add the amount for that category
            if record[2] in tableA:
                tableA[record[2]] += record[3] 
            else:
                tableA[record[2]] = record[3]

    return tableA


# FUNCTION - Create Table B
# Amounts for each month

def createTableB():
    # Obtain 2D array with transaction records
    recordsTable = obtainRecords()
    tableB = {}

    for record in recordsTable:
        if record[0] != 'DATE': # Don't count the heading 

            # Create a string with the year and month
            recordDate = str(record[0])
            recordMonth = str(recordDate[0:4]+'-'+str(recordDate[5:7]))

            # Add the amount for that month
            if recordMonth in tableB:
                tableB[recordMonth] += record[3]
            else:
                tableB[recordMonth] = record[3]

    return tableB


# FUNCTION - Create Table C
# Amounts for each category for each month

def createTableC():
# Obtain 2D array with transaction records
    recordsTable = obtainRecords()
    tableC = {}

    for record in recordsTable:
        if record[0] != 'DATE': # Don't count the heading
            
            # Create a dictionary for the year and month

            # Create a string with the year and month
            recordDate = str(record[0])
            recordMonth = str(recordDate[0:4]+'-'+str(recordDate[5:7]))

            # Create a nested dictionary for that month
            if recordMonth not in tableC:
                tableC[recordMonth] = {}

            # Add the amount for the category for that month into the nested dictionary
            if record[2] in tableC[recordMonth]:
                tableC[recordMonth][record[2]] += record[3]
            else:
                tableC[recordMonth][record[2]] = record[3]

    return tableC


# MAIN CODE

# Load Excel file with personal expense transactions
petFileName = "Money_Book.xlsx"
petWorkBook = load_workbook(petFileName)
pet = petWorkBook.active

# Create home window
home = tk.Tk()
home.title("Money Book")
home.geometry("600x400")

# Variables

# Calendar
calendar = {
    '01': 'January',
    '02': 'February',
    '03': 'March',
    '04': 'April',
    '05': 'May',
    '06': 'June',
    '07': 'July',
    '08': 'August',
    '09': 'September',
    '10': 'October',
    '11': 'November',
    '12': 'December'
}

# Extract today's date
today = dt.datetime.today() 

# Categories List
categoriesList = ('Food', 'Housing', 'Leisure', 'Maintenance', 'Goods', 'Travel', 'Bills', 'Other')

# Style Sheet
font = 'Arial'
heading1 = 24
heading2 = 18
heading3 = 16
textSize = 10

# HOME SCREEN

# Application Information
tk.Label(home, text="Money Book", font=(font, heading1)).pack()
tk.Label(home, text="Personal Expense Tracker", font=(font, heading2)).pack()
tk.Label(home, text="Version 1.0", font=(font, heading3)).pack()
tk.Label(home, text="March 2026", font=(font, heading3)).pack()

# Options Menu
optionsBar = Menu(home)
home.config(menu=optionsBar)
options = Menu(optionsBar, tearoff=0)
optionsBar.add_cascade(label="Options", font=(font, textSize),  menu=options)
options.add_command(label="Entry Form", font=(font, textSize), command=entryForm)
options.add_command(label="Categories", font=(font, textSize), command=categories)
options.add_command(label="Monthly Balances", font=(font, textSize), command=monthlyBalances)
options.add_command(label="Statistics", font=(font, textSize), command=statistics)
options.add_command(label="Exit", font=(font, textSize), command=exit)
home.config(menu=optionsBar)

home.mainloop()
